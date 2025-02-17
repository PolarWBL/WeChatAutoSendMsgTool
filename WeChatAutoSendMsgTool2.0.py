import os
import time
import uiautomation as auto  # 导入uiautomation库,用于自动化控制应用程序界面
import win32clipboard  # 导入win32clipboard库,用于剪贴板操作
import keyboard # 导入keyboard库,用于模拟键盘操作
from ctypes import *  # 导入ctypes库,用于创建C语言结构体
from openpyxl import load_workbook # 导入openpyxl库,用于读取Excel文件

MAIN_PATH = os.getcwd()+'\\' # 主目录路径

# 定义DROPFILES结构体,代表Windows的文件拖放数据结构
class DROPFILES(Structure):
    _fields_ = [
        ("pFiles", c_uint),  # 文件信息结构的大小
        ("x", c_long),       # 鼠标位置X坐标（拖放时的位置）
        ("y", c_long),       # 鼠标位置Y坐标（拖放时的位置）
        ("fNC", c_int),      # 网络计算机的标志（用于网络路径）
        ("fWide", c_bool),   # 是否使用宽字符（True为宽字符,False为ANSI字符）
    ]

# 创建DROPFILES实例并初始化字段
pDropFiles = DROPFILES()
pDropFiles.pFiles = sizeof(DROPFILES)  # 设置结构体的大小
pDropFiles.fWide = True  # 设置为宽字符（Unicode）
matedata = bytes(pDropFiles)  # 将DROPFILES结构体转化为字节流

# 设置微信窗口控制对象
wechatWindow = auto.WindowControl(searchDepth=1, ClassName='mmui::MainWindow')  # 寻找微信主窗口
search = wechatWindow.EditControl(ClassName='mmui::XLineEdit')  # 定位到搜索框
edit = wechatWindow.EditControl(ClassName='mmui::ChatInputField')  # 定位到输入框

# 将文件路径列表设置到剪贴板
def setClipboardFiles(paths):
    """
    将文件路径列表放入剪贴板
    :param paths: 文件路径列表
    """
    # 将路径列表转化为以\0分隔的字符串,并将所有的路径分隔符替换为\\
    files = ("\0".join(paths)).replace("/", "\\")

    # 将路径字符串编码为Unicode（U16格式）,去掉前两个字节,并添加结束符
    data = files.encode("U16")[2:] + b"\0\0"
    
    # 打开剪贴板
    win32clipboard.OpenClipboard()
    try:
        # 清空剪贴板
        win32clipboard.EmptyClipboard()
        
        # 将文件路径数据以CF_HDROP格式写入剪贴板,包含文件路径的元数据
        win32clipboard.SetClipboardData(
            win32clipboard.CF_HDROP, matedata + data
        )
    finally:
        # 关闭剪贴板
        win32clipboard.CloseClipboard()

# 只将单个文件路径放入剪贴板
def setClipboardFile(file):
    """
    将单个文件路径放入剪贴板
    :param file: 文件路径
    """
    setClipboardFiles([file])  # 调用setClipboardFiles函数

# 从剪贴板读取文件路径列表
def readClipboardFilePaths():
    """
    从剪贴板读取文件路径
    :return: 返回文件路径列表
    """
    # 打开剪贴板
    win32clipboard.OpenClipboard()
    paths = None
    try:
        # 获取剪贴板中的文件路径数据,CF_HDROP格式
        return win32clipboard.GetClipboardData(win32clipboard.CF_HDROP)
    finally:
        # 关闭剪贴板
        win32clipboard.CloseClipboard()

# 通过联系人名称选择会话
def selectSessionFromName(name, wait_time=0.5):
    """
    选择指定名称的聊天会话
    :param name: 好友或群聊的名称
    :param wait_time: 等待时间,默认0.5秒
    """
    search.Click()  # 点击搜索框
    auto.SetClipboardText(name)  # 将好友名称复制到剪贴板
    search.SendKeys('{Ctrl}a')  # 粘贴好友名称到输入框
    search.SendKeys('{Ctrl}v')  # 粘贴好友名称到输入框
    time.sleep(wait_time)  # 等待微信索引搜索结果
    search.SendKeys("{Enter}")  # 按回车进入会话

# 发送消息函数,支持文本、图片和文件
def send_msg(content, msg_type=1):
    """
    发送消息
    :param content: 消息内容
    :param msg_type: 消息类型,1为文本,2为文件
    """
    if msg_type == 1:
        # 处理文本消息
        auto.SetClipboardText(content)  # 将消息内容设置到剪贴板
    elif msg_type == 2:
        # 处理文件消息
        setClipboardFile(content)  # 将文件路径设置到剪贴板
    edit.SendKeys('{Ctrl}v')  # 粘贴消息内容（文本、图片或文件）

# 从Excel文件中读取群发信息列表并发送
def readInfo(strfile):
    try:
        print("========================================")
        print(f"【开始读取数据】正在读取 {strfile} 中的信息...")
        sheet = load_workbook(strfile).worksheets[0]
        listData = []
        index = 1
        for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始读取
            print(f"【读取中】正在读取第 {index+1} 行数据")
            
            # 校验姓名
            name = str(row[0]) if row[0] else ''.strip()
            if not name:
                print(f"【警告】第 {index+1} 行 [*搜索] 内容为空")
                index += 1
                continue  # 如果姓名为空则跳过

            # 拼接消息内容
            content = ''.join([str(row[i]) if row[i] else '' for i in range(1, 4)])  # 使用 ''.join() 进行拼接

            # 校验文件路径
            file_path = str(row[4]) if row[4] else ''.strip()  # 处理文件路径为空的情况
            if file_path :
                file_path = MAIN_PATH + 'files\\' + file_path
                if not os.path.exists(file_path):
                    print(f"【警告】第 {index+1} 行附件 {file_path} 不存在")
                    file_path = ''
            else:
                file_path = ''

            # 将数据添加到listData中
            info = {
                'name': name,
                'content': content,
                'file_path': file_path,
            }
            listData.append(info)
            index += 1
        print(f"【数据读取完毕】")
        return listData
    except Exception as e:
        print(f"【数据读取失败】: {e}")
        return []    

# 发送消息列表
def messagesend(listData):
    wechatWindow.SetActive()  # 激活微信窗口
    print(f"【发送给】{listData['name']}")
    selectSessionFromName(listData['name']) # 选择要发送的人
    if listData['content']:  # 如果消息内容不为空
        print(f"【消息内容】\n{listData['content']}")
        send_msg(listData['content'], 1)  # 发送文字消息     
    if listData['file_path']:  # 如果文件路径不为空
        print(f"【发送附件】\n{listData['file_path']}")
        send_msg(listData['file_path'], 2) # 发送文件
    time.sleep(1)  # 等待1秒
    edit.SendKeys('{Alt}s')  # 发送消息

# 主程序
if __name__ == '__main__':
    print("========================================")
    print("【提示】请确保微信已经登录并打开主界面,输入'y'键继续,输入其他内容退出程序...");
    while True:
        if input().lower() == 'y': # 等待按下'y'键
            file= 'WeChatAutoSendMsgList.xlsx'
            listData = readInfo(MAIN_PATH+file)
            if not listData:
                print("【警告】表格文件无数据或读取失败,请检查文件内容...")
            else:
                for index, data in enumerate(listData):
                    print("========================================")
                    print(f"【执行中】读取第 {index+1} 个发送对象的信息")
                    messagesend(data)
                    index += 1
                print("========================================")
                print(f"【执行结果】向 {len(listData)} 个对象发送消息.")
            print("【执行结束】输入'y'键再次发送,输入其他内容退出程序...")
        else:
            break